"""Chargrill award-pay engine — ported from the standalone Payroll System
(run_payroll.py) so the COGS app can compute labour cost from a weekly Tanda
shift CSV.

Award reference: Fast Food Industry Award 2020 (MA000003).

Decoupled from files/CLI: setup + CSV come in as bytes, results come back as
plain dicts, and the Excel report is built into memory. The calculation
functions (process_employee / calculate_pay / get_rates_for_employee) are kept
identical to the original tool so the numbers match exactly.
"""
import io
import re
import pandas as pd
from datetime import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DAILY_OT_THRESHOLD = 11.0   # single day hours that trigger daily OT
WEEKLY_OT_THRESHOLD = 38.0  # weekly hours that trigger weekly OT (FT default)
LAUNDRY_PER_SHIFT = 1.25    # laundry allowance per shift (Casual & Part-Time only)

# ── STYLES (Excel report) ────────────────────────────────────────────────────
HDR_BG = '1F3864'; HDR_FG = 'FFFFFF'
SUB_BG = '2E75B6'; SUB_FG = 'FFFFFF'
TOPUP_BG = 'FF6600'
ZERO_BG = 'E2EFDA'
OT_BG = 'FFF3CD'
PH_BG = 'FCE4D6'
SAT_BG = 'E8F4FD'
SUN_BG = 'EBF3FB'
CAS_BG = 'F0F7EC'
GREY_BG = 'F2F2F2'
TOTAL_BG = 'D9D9D9'
FOH_BG = 'DAEEF3'
BOH_BG = 'FEF9E7'
BLUE_FG = '0000FF'
BLACK_FG = '000000'
GREY_FG = '595959'
RED_FG = 'FF0000'
CUR_FMT = '#,##0.00'
HRS_FMT = '0.00'
PCT_FMT = '0.0%'


def tb(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def sc(cell, value=None, bold=False, size=9, fg=BLACK_FG, bg=None,
       halign='left', valign='center', wrap=False, fmt=None, bdr=True, italic=False):
    if value is not None:
        cell.value = value
    cell.font = Font(name='Arial', bold=bold, size=size, color=fg, italic=italic)
    if bg:
        cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    if bdr:
        cell.border = tb()


# ── SETUP + CSV LOADING (from bytes) ─────────────────────────────────────────

def normalise_name(name):
    """Lowercase, collapse whitespace for fuzzy matching."""
    return re.sub(r'\s+', ' ', str(name).strip().lower())


def load_setup_from_bytes(xlsx_bytes):
    """Parse Payroll Setup.xlsx (in memory) -> (emp_df, rates, public_holidays)."""
    buf = io.BytesIO(xlsx_bytes)
    emp_df = pd.read_excel(buf, sheet_name='EMPLOYEES', skiprows=3)
    emp_df.columns = [str(c).replace('\n', ' ').strip() for c in emp_df.columns]
    emp_df = emp_df.dropna(subset=['Employee Name (Display)'])
    emp_df['_csv_key'] = emp_df['CSV Name (Exact match in Tanda CSV)'].fillna(
        emp_df['Employee Name (Display)']).apply(normalise_name)

    buf.seek(0)
    rates_df = pd.read_excel(buf, sheet_name='AWARD RATES', skiprows=4, header=0)
    rates_df.columns = [str(c).strip() for c in rates_df.columns]

    buf.seek(0)
    ph_df = pd.read_excel(buf, sheet_name='PUBLIC HOLIDAYS', skiprows=2, header=0)
    ph_df.columns = [str(c).strip() for c in ph_df.columns]
    ph_df = ph_df.dropna(subset=['Date'])
    public_holidays = set(pd.to_datetime(ph_df['Date']).dt.date)

    rates = _parse_rates(rates_df)
    return emp_df, rates, public_holidays


def _parse_rates(df):
    """Extract rate values from the Award Rates sheet into a structured dict."""
    defaults = {
        'casual': {
            'weekday': 33.19, 'saturday': 39.83, 'sunday': 39.83,
            'ph_worked': 66.38, 'ph_unworked': 33.19, 'ph_ot': 73.01,
            'daily_ot_1': 46.46, 'daily_ot_2': 59.74,
            'weekly_ot_1': 46.46, 'weekly_ot_2': 59.74,
            'late_night': 35.84, 'laundry': 6.25,
        },
        'permanent': {
            'weekday': 26.55, 'saturday': 33.19, 'sunday': 33.19,
            'ph_worked': 59.74, 'ph_unworked': 26.55, 'ph_ot': 66.38,
            'daily_ot_1': 39.83, 'daily_ot_2': 53.10,
            'weekly_ot_1': 39.83, 'weekly_ot_2': 53.10,
            'late_night': 28.67, 'laundry': 0.0,
        },
    }
    keywords = {
        'Ordinary Weekday': ('weekday', (1, 2)),
        'Saturday Rate': ('saturday', (1, 2)),
        'Sunday Rate': ('sunday', (1, 2)),
        'Public Holiday — Worked': ('ph_worked', (1, 2)),
        'Public Holiday — Unworked': ('ph_unworked', (1, 2)),
        'Public Holiday OT': ('ph_ot', (1, 2)),
        'Daily OT — First 2': ('daily_ot_1', (1, 2)),
        'Daily OT — After 2': ('daily_ot_2', (1, 2)),
        'Weekly OT — First 2': ('weekly_ot_1', (1, 2)),
        'Weekly OT — After 2': ('weekly_ot_2', (1, 2)),
        'Late Night': ('late_night', (1, 2)),
        'Laundry': ('laundry', (1, 2)),
    }
    try:
        for _, row in df.iterrows():
            component = str(row.iloc[0])
            for kw, (rate_key, (ci, pi)) in keywords.items():
                if kw.lower() in component.lower():
                    try:
                        cas_val = float(row.iloc[ci])
                        if not pd.isna(cas_val):
                            defaults['casual'][rate_key] = cas_val
                    except Exception:
                        pass
                    try:
                        perm_val = float(row.iloc[pi])
                        if not pd.isna(perm_val):
                            defaults['permanent'][rate_key] = perm_val
                    except Exception:
                        pass
    except Exception:
        pass
    return defaults


def find_employee(csv_name, emp_df):
    """Match a CSV name to an employee record."""
    key = normalise_name(csv_name)
    match = emp_df[emp_df['_csv_key'] == key]
    if len(match) == 0:
        for _, row in emp_df.iterrows():
            if key in row['_csv_key'] or row['_csv_key'] in key:
                return row
        return None
    return match.iloc[0]


def load_csv_from_bytes(csv_bytes):
    """Parse a weekly Tanda shift CSV (in memory) into a cleaned DataFrame."""
    df = pd.read_csv(io.BytesIO(csv_bytes))
    df.columns = [c.strip() for c in df.columns]
    df = df[~df['Name'].str.strip().str.startswith('Total', na=True)]
    df = df.dropna(subset=['Name'])
    df['Shift Length'] = pd.to_numeric(df['Shift Length'], errors='coerce').fillna(0)
    df = df[df['Shift Length'] > 0]
    df = df.dropna(subset=['Shift Start Time', 'Shift End Time'])
    df = df[df['Shift Start Time'].astype(str).str.strip() != '']
    df = df[df['Shift End Time'].astype(str).str.strip() != '']
    df['Date'] = pd.to_datetime(df['Date'])
    df['_name_key'] = df['Name'].apply(normalise_name)
    return df


# ── TIME HELPERS ─────────────────────────────────────────────────────────────

def parse_hhmm(s):
    try:
        parts = str(s).strip().split(':')
        return time(int(parts[0]) % 24, int(parts[1]))
    except Exception:
        return None


def late_night_hrs(start_s, end_s, cutoff_h=22):
    """Hours worked after cutoff_h within a single shift."""
    st = parse_hhmm(start_s)
    et = parse_hhmm(end_s)
    if not st or not et:
        return 0.0
    s_m = st.hour * 60 + st.minute
    e_m = et.hour * 60 + et.minute
    c_m = cutoff_h * 60
    mn_m = 24 * 60
    if e_m < s_m:
        e_m += mn_m
    late_s = max(s_m, c_m)
    late_e = min(e_m, mn_m)
    return max(0.0, (late_e - late_s) / 60.0)


def day_type(d, public_holidays):
    if isinstance(d, pd.Timestamp):
        d = d.date()
    if d in public_holidays:
        return 'ph'
    wd = d.weekday()  # 0=Mon … 6=Sun
    if wd == 5:
        return 'sat'
    if wd == 6:
        return 'sun'
    return 'wd'


# ── AWARD CALCULATION ────────────────────────────────────────────────────────

def process_employee(emp_shifts, public_holidays, emp_type, min_weekly_hrs=38):
    """Hours breakdown + pay calculation for one employee."""
    hrs = dict(wd=0.0, sat=0.0, sun=0.0, sun_ot=0.0, ph=0.0,
               late_night=0.0, daily_ot1=0.0, daily_ot2=0.0,
               ph_daily_ot=0.0, weekly_ot1=0.0, weekly_ot2=0.0, total=0.0)
    day_rows = []

    for dt_val, day_df in emp_shifts.groupby('Date'):
        day_total = day_df['Shift Length'].sum()
        hrs['total'] += day_total
        dtype = day_type(dt_val, public_holidays)

        if day_total > DAILY_OT_THRESHOLD:
            excess = day_total - DAILY_OT_THRESHOLD
            ot1 = min(excess, 2.0)
            ot2 = max(0.0, excess - 2.0)
            day_ord = DAILY_OT_THRESHOLD
            if dtype == 'ph':
                hrs['ph_daily_ot'] += ot1 + ot2
            else:
                hrs['daily_ot1'] += ot1
                hrs['daily_ot2'] += ot2
        else:
            day_ord = day_total
            ot1 = ot2 = 0.0

        ln = sum(late_night_hrs(r['Shift Start Time'], r['Shift End Time'])
                 for _, r in day_df.iterrows())
        hrs['late_night'] += ln

        # All ordinary hours are paid at the day's penalty rate, Sundays included.
        # Overtime is handled above via the daily (>11h) / weekly (>38h) thresholds —
        # there is no separate "only the first 2 Sunday hours are ordinary" rule.
        sun_ot_hrs = 0.0
        hrs[dtype] += day_ord

        day_rows.append({
            'date': dt_val,
            'day_name': pd.Timestamp(dt_val).day_name()[:3],
            'day_type': dtype,
            'shifts': len(day_df),
            'total_hrs': day_total,
            'ordinary_hrs': day_ord,
            'sun_ot_hrs': sun_ot_hrs,
            'daily_ot1': ot1,
            'daily_ot2': ot2,
            'late_night': ln,
        })

    wk_threshold = min_weekly_hrs if (isinstance(min_weekly_hrs, (int, float))
                                      and min_weekly_hrs > 0) else WEEKLY_OT_THRESHOLD
    total_ord = hrs['wd'] + hrs['sat'] + hrs['sun'] + hrs['sun_ot'] + hrs['ph']

    if total_ord > wk_threshold:
        wk_ot = total_ord - wk_threshold
        hrs['weekly_ot1'] = min(wk_ot, 2.0)
        hrs['weekly_ot2'] = max(0.0, wk_ot - 2.0)
        reduce = wk_ot
        wd_take = min(reduce, hrs['wd'])
        hrs['wd'] -= wd_take
        reduce -= wd_take
        if reduce > 0:
            sat_take = min(reduce, hrs['sat'])
            hrs['sat'] -= sat_take

    return hrs, day_rows


def calculate_pay(hrs, rates):
    """Convert hours dict → pay amounts dict."""
    r = rates
    late_loading = max(0.0, r['late_night'] - r['weekday'])
    pay = {
        'wd':        hrs['wd']          * r['weekday'],
        'sat':       hrs['sat']         * r['saturday'],
        'sun':       hrs['sun']         * r['sunday'],
        'sun_ot':    hrs['sun_ot']      * r['daily_ot_2'],  # Sunday OT (>2h) at double-time
        'ph':        hrs['ph']          * r['ph_worked'],
        'ph_ot':     hrs['ph_daily_ot'] * r['ph_ot'],
        'daily_ot1': hrs['daily_ot1']   * r['daily_ot_1'],
        'daily_ot2': hrs['daily_ot2']   * r['daily_ot_2'],
        'wk_ot1':    hrs['weekly_ot1']  * r['weekly_ot_1'],
        'wk_ot2':    hrs['weekly_ot2']  * r['weekly_ot_2'],
        'late_night': hrs['late_night'] * late_loading,
        'laundry':   r.get('laundry', 0.0),
    }
    pay['total'] = sum(pay.values())
    return pay


def get_emp_type_key(emp_type_str):
    et = str(emp_type_str).strip().lower()
    if 'casual' in et:
        return 'Casual'
    if 'part' in et:
        return 'Part-Time'
    return 'Full-Time'


def _get_col(row_dict, *keys):
    """Return first non-None value from a list of possible column names."""
    for k in keys:
        v = row_dict.get(k)
        if v is not None and str(v).strip() not in ('', 'nan', 'None'):
            return v
    return None


def get_rates_for_employee(emp_row, all_rates):
    """Return the appropriate rates dict based on employment type/level/age."""
    emp_type_str = _get_col(emp_row, 'Employment Type', 'Employment\nType') or 'Full-Time'
    emp_type = get_emp_type_key(emp_type_str)
    base_rates = all_rates['casual'] if emp_type == 'Casual' else all_rates['permanent']

    level_mult = {'1': 1.000, '2': 1.052, '3': 1.104}
    level = str(_get_col(emp_row, 'Award Level', 'Award\nLevel') or '1').strip()
    lm = level_mult.get(level, 1.0)

    age_mult = {'adult': 1.000, 'u21': 0.893, 'u20': 0.788, 'u19': 0.683,
                'u18': 0.578, 'u17': 0.473, 'u16': 0.368}
    age_cat = str(_get_col(emp_row, 'Age Category', 'Age\nCategory') or 'Adult').strip().lower()
    am = age_mult.get(age_cat, 1.0)

    multiplier = lm * am
    rates = {}
    for k, v in base_rates.items():
        rates[k] = v if k == 'laundry' else round(v * multiplier, 4)

    rates.setdefault('daily_ot_1', rates.get('daily_ot1', 0))
    rates.setdefault('daily_ot_2', rates.get('daily_ot2', 0))
    rates.setdefault('weekly_ot_1', rates.get('weekly_ot1', 0))
    rates.setdefault('weekly_ot_2', rates.get('weekly_ot2', 0))
    return rates


# ── ORCHESTRATION ────────────────────────────────────────────────────────────

def process_shifts(shift_df, emp_df, all_rates, public_holidays):
    """Run the award calc across every employee in the shift CSV.

    Returns a dict:
      results       list of per-employee dicts (name, emp_type, section, hrs,
                    pay, rates, award_pay, flat_pay, topup, gross, day_rows)
      week_ending   pandas Timestamp (max shift date)
      unmatched     CSV names not found in the setup sheet
      total_gross / total_topup / total_hours
    """
    week_ending = shift_df['Date'].max()
    all_results, unmatched = [], []

    for csv_key in sorted(shift_df['_name_key'].unique()):
        emp_row = find_employee(csv_key, emp_df)
        emp_shifts = shift_df[shift_df['_name_key'] == csv_key].copy()

        if emp_row is None:
            unmatched.append(csv_key)
            display_name = emp_shifts['Name'].iloc[0]
            emp_row_dict = {
                'Employee Name (Display)': display_name,
                'Employment Type': 'Full-Time', 'Award Level': '1',
                'Age Category': 'Adult', 'Flat Hourly Rate ($)': 0,
                'Section (FOH/BOH)': '', 'Min Weekly Hours': 38,
            }
        else:
            emp_row_dict = emp_row.to_dict()

        emp_type = get_emp_type_key(
            _get_col(emp_row_dict, 'Employment Type', 'Employment\nType') or 'Full-Time')
        flat_rate_raw = _get_col(emp_row_dict, 'Flat Hourly Rate ($)', 'Flat Hourly\nRate ($)')
        flat_rate = float(flat_rate_raw) if flat_rate_raw not in (None, '', 'nan') else 0.0
        min_weekly_raw = _get_col(emp_row_dict, 'Min Weekly Hours', 'Min Weekly\nHours')
        min_weekly = float(min_weekly_raw) if min_weekly_raw not in (None, '', 'nan') else 38.0
        section = str(_get_col(emp_row_dict, 'Section (FOH/BOH)', 'Section\n(FOH/BOH)') or '').strip()

        rates = get_rates_for_employee(emp_row_dict, all_rates)
        hrs, day_rows = process_employee(emp_shifts, public_holidays, emp_type, min_weekly)
        pay = calculate_pay(hrs, rates)

        # Laundry allowance: $1.25 per shift for Casual & Part-Time (else none).
        n_shifts = sum(int(d.get('shifts', 0) or 0) for d in day_rows)
        laundry = round(LAUNDRY_PER_SHIFT * n_shifts, 2) if emp_type in ('Casual', 'Part-Time') else 0.0
        pay['total'] = round(pay['total'] - pay.get('laundry', 0.0) + laundry, 4)
        pay['laundry'] = laundry

        flat_pay = hrs['total'] * flat_rate
        award_pay = pay['total']
        topup = max(0.0, award_pay - flat_pay) if emp_type != 'Casual' else 0.0
        gross = flat_pay + topup if emp_type != 'Casual' else award_pay

        display_name = (_get_col(emp_row_dict, 'Employee Name (Display)', 'Employee Name')
                        or emp_shifts['Name'].iloc[0])

        all_results.append({
            'name': display_name, 'emp_type': emp_type, 'section': str(section).strip(),
            'flat_rate': flat_rate, 'hrs': hrs, 'pay': pay, 'rates': rates,
            'award_pay': round(award_pay, 2), 'flat_pay': round(flat_pay, 2),
            'topup': round(topup, 2), 'gross': round(gross, 2), 'day_rows': day_rows,
        })

    all_results.sort(key=lambda r: (0 if r['emp_type'] != 'Casual' else 1, r['name']))
    total_gross = round(sum(r['gross'] for r in all_results), 2)
    total_topup = round(sum(r['topup'] for r in all_results), 2)
    total_hours = round(sum(r['hrs']['total'] for r in all_results), 2)
    return {
        'results': all_results, 'week_ending': week_ending, 'unmatched': unmatched,
        'total_gross': total_gross, 'total_topup': total_topup, 'total_hours': total_hours,
    }


def process_shift_csv(csv_bytes, setup_bytes):
    """Convenience: setup + CSV bytes -> full results dict (see process_shifts)."""
    emp_df, rates, public_holidays = load_setup_from_bytes(setup_bytes)
    shift_df = load_csv_from_bytes(csv_bytes)
    return process_shifts(shift_df, emp_df, rates, public_holidays)


def apply_leave(out, leave):
    """Add annual/sick leave (paid at the flat rate) to FT/PT employees and refresh
    the totals. `leave` = {employee_name: {'al': hours, 'sl': hours}}. Idempotent —
    each call recomputes gross from base (flat pay + top-up), so it can run on every
    rerun. Casuals accrue no paid leave and are left unchanged."""
    for r in out["results"]:
        if r["emp_type"] == "Casual":
            r["al_hrs"] = r["sl_hrs"] = r["leave_pay"] = 0.0
            continue
        lv = leave.get(r["name"], {}) or {}
        al = float(lv.get("al", 0) or 0)
        sl = float(lv.get("sl", 0) or 0)
        r["al_hrs"], r["sl_hrs"] = al, sl
        r["leave_pay"] = round((al + sl) * r["flat_rate"], 2)
        r["gross"] = round(r["flat_pay"] + r["topup"] + r["leave_pay"], 2)
    out["total_gross"] = round(sum(r["gross"] for r in out["results"]), 2)
    out["total_leave"] = round(sum(r.get("leave_pay", 0) for r in out["results"]), 2)
    return out


# ── EXCEL REPORT (built in memory) ───────────────────────────────────────────

def build_workbook(all_results, week_ending) -> bytes:
    """Build the same 4-sheet Excel report as the standalone tool, into bytes."""
    wb = openpyxl.Workbook()
    ws_pay = wb.active
    ws_pay.title = 'PAYROLL SUMMARY'
    ws_cas = wb.create_sheet('CASUAL DETAIL')
    ws_lab = wb.create_sheet('LABOUR COSTS')
    ws_day = wb.create_sheet('DAILY BREAKDOWN')

    week_str = pd.Timestamp(week_ending).strftime('%d %b %Y')
    _payroll_summary(ws_pay, all_results, week_str)
    _casual_detail(ws_cas, all_results, week_str)
    _labour_costs(ws_lab, all_results, week_str)
    _daily_breakdown(ws_day, all_results, week_str)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _payroll_summary(ws, results, week_str):
    ws.sheet_view.showGridLines = False
    perm_results = [r for r in results if r['emp_type'] != 'Casual']
    cas_results = [r for r in results if r['emp_type'] == 'Casual']
    INPUT_BG = 'FFF2CC'

    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    col_widths = [24, 10, 8, 10, 10, 8, 8, 10, 12, 10, 10, 10, 12]
    for col, w in zip(cols, col_widths):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:M1')
    ws.row_dimensions[1].height = 32
    sc(ws['A1'], f'PAYROLL SUMMARY  —  Week Ending {week_str}',
       bold=True, size=14, fg=HDR_FG, bg=HDR_BG, halign='center', valign='center')

    row = 3
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f'A{row}:M{row}')
    sc(ws[f'A{row}'], 'FULL-TIME & PART-TIME EMPLOYEES  —  Flat Rate vs Award Comparison',
       bold=True, size=10, fg=HDR_FG, bg=SUB_BG, halign='left')

    row += 1
    ws.row_dimensions[row].height = 30
    perm_hdrs = ['Employee', 'Type', 'Section', 'Flat Rate\n($)', 'Worked\nHrs',
                 'SL Hrs\n(input)', 'AL Hrs\n(input)', 'Total\nHrs',
                 'Flat Pay', 'SL Pay', 'AL Pay', 'Top Up', 'GROSS PAY']
    for i, hdr in enumerate(perm_hdrs):
        sc(ws[f'{cols[i]}{row}'], hdr, bold=True, size=9, fg=HDR_FG, bg='365F91',
           halign='center', wrap=True)

    perm_data_start = row + 1
    for r_data in perm_results:
        row += 1
        ws.row_dimensions[row].height = 16
        has_topup = r_data['topup'] > 0.01
        bg = TOPUP_BG if has_topup else ZERO_BG
        sc(ws[f'A{row}'], r_data['name'], size=9, fg=BLUE_FG, halign='left')
        sc(ws[f'B{row}'], r_data['emp_type'], size=9, halign='center')
        sc(ws[f'C{row}'], r_data['section'], size=9, halign='center')
        sc(ws[f'D{row}'], r_data['flat_rate'], size=9, bg=bg, fmt=CUR_FMT, halign='center', fg=BLUE_FG)
        sc(ws[f'E{row}'], r_data['hrs']['total'], size=9, bg=bg, fmt=HRS_FMT, halign='center')
        sc(ws[f'F{row}'], r_data.get('sl_hrs', 0) or 0, size=9, bg=INPUT_BG, fmt=HRS_FMT, halign='center')
        sc(ws[f'G{row}'], r_data.get('al_hrs', 0) or 0, size=9, bg=INPUT_BG, fmt=HRS_FMT, halign='center')
        sc(ws[f'H{row}'], f'=E{row}+F{row}+G{row}', size=9, bg=bg, fmt=HRS_FMT, halign='center')
        sc(ws[f'I{row}'], r_data['flat_pay'], size=9, bg=bg, fmt=CUR_FMT, halign='center')
        sc(ws[f'J{row}'], f'=D{row}*F{row}', size=9, bg=bg, fmt=CUR_FMT, halign='center')
        sc(ws[f'K{row}'], f'=D{row}*G{row}', size=9, bg=bg, fmt=CUR_FMT, halign='center')
        sc(ws[f'L{row}'], r_data['topup'], size=9, bg=bg, fmt=CUR_FMT, halign='center',
           fg=RED_FG if has_topup else BLACK_FG, bold=has_topup)
        sc(ws[f'M{row}'], f'=I{row}+J{row}+K{row}+L{row}', size=9, bg=bg, fmt=CUR_FMT,
           halign='center', bold=True)

    perm_data_end = row
    row += 1
    perm_sub_row = row
    ws.row_dimensions[row].height = 18
    sc(ws[f'A{row}'], 'SUBTOTAL', bold=True, size=9, bg=TOTAL_BG)
    for c in 'BCD':
        sc(ws[f'{c}{row}'], '', bold=True, size=9, bg=TOTAL_BG)
    for c in 'EFGH':
        sc(ws[f'{c}{row}'], f'=SUM({c}{perm_data_start}:{c}{perm_data_end})',
           bold=True, size=9, bg=TOTAL_BG, fmt=HRS_FMT, halign='center')
    for c in 'IJKLM':
        sc(ws[f'{c}{row}'], f'=SUM({c}{perm_data_start}:{c}{perm_data_end})',
           bold=True, size=9, bg=TOTAL_BG, fmt=CUR_FMT, halign='center',
           fg=RED_FG if c == 'L' else BLACK_FG)

    row += 2
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f'A{row}:M{row}')
    sc(ws[f'A{row}'], 'CASUAL EMPLOYEES  —  Award Pay (see CASUAL DETAIL sheet for full breakdown)',
       bold=True, size=10, fg=HDR_FG, bg=SUB_BG, halign='left')

    row += 1
    ws.row_dimensions[row].height = 22
    cas_hdrs = ['Employee', 'Type', 'Section', '', 'Hours', '', '', '', 'Award Pay', '', '', '', 'GROSS PAY']
    for i, hdr in enumerate(cas_hdrs):
        sc(ws[f'{cols[i]}{row}'], hdr, bold=True, size=9, fg=HDR_FG, bg='375623',
           halign='center', wrap=True)

    cas_data_start = row + 1
    for r_data in cas_results:
        row += 1
        ws.row_dimensions[row].height = 16
        for c in cols:
            sc(ws[f'{c}{row}'], '', size=9, bg=CAS_BG)
        sc(ws[f'A{row}'], r_data['name'], size=9, bg=CAS_BG, halign='left')
        sc(ws[f'B{row}'], 'Casual', size=9, bg=CAS_BG, halign='center')
        sc(ws[f'C{row}'], r_data['section'], size=9, bg=CAS_BG, halign='center')
        sc(ws[f'E{row}'], r_data['hrs']['total'], size=9, bg=CAS_BG, fmt=HRS_FMT, halign='center')
        sc(ws[f'I{row}'], r_data['award_pay'], size=9, bg=CAS_BG, fmt=CUR_FMT, halign='center')
        sc(ws[f'M{row}'], r_data['award_pay'], size=9, bg=CAS_BG, fmt=CUR_FMT, halign='center', bold=True)

    cas_data_end = row
    row += 1
    cas_sub_row = row
    ws.row_dimensions[row].height = 18
    sc(ws[f'A{row}'], 'SUBTOTAL', bold=True, size=9, bg=TOTAL_BG)
    for c in cols[1:]:
        sc(ws[f'{c}{row}'], '', bold=True, size=9, bg=TOTAL_BG)
    sc(ws[f'E{row}'], f'=SUM(E{cas_data_start}:E{cas_data_end})', bold=True, size=9, bg=TOTAL_BG, fmt=HRS_FMT, halign='center')
    sc(ws[f'I{row}'], f'=SUM(I{cas_data_start}:I{cas_data_end})', bold=True, size=9, bg=TOTAL_BG, fmt=CUR_FMT, halign='center')
    sc(ws[f'M{row}'], f'=SUM(M{cas_data_start}:M{cas_data_end})', bold=True, size=9, bg=TOTAL_BG, fmt=CUR_FMT, halign='center')

    row += 2
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f'A{row}:D{row}')
    sc(ws[f'A{row}'], 'GRAND TOTAL PAYROLL COST', bold=True, size=11, fg=HDR_FG, bg=HDR_BG)
    for c in cols[4:]:
        sc(ws[f'{c}{row}'], '', bold=True, size=11, bg=HDR_BG)
    sc(ws[f'E{row}'], f'=E{perm_sub_row}+E{cas_sub_row}', bold=True, size=11, fg=HDR_FG, bg=HDR_BG, fmt=HRS_FMT, halign='center')
    sc(ws[f'L{row}'], f'=L{perm_sub_row}', bold=True, size=11, fg=HDR_FG, bg=HDR_BG, fmt=CUR_FMT, halign='center')
    sc(ws[f'M{row}'], f'=M{perm_sub_row}+M{cas_sub_row}', bold=True, size=11, fg=HDR_FG, bg=HDR_BG, fmt=CUR_FMT, halign='center')

    row += 1
    ws.row_dimensions[row].height = 30
    ws.merge_cells(f'A{row}:M{row}')
    sc(ws[f'A{row}'],
       'TOP UP amounts (orange rows) must be added to the employee\'s regular flat-rate pay.  '
       'Yellow SL/AL Hrs cells: type hours manually — Total Hrs, SL Pay, AL Pay, and GROSS PAY update automatically.  '
       'Always verify award rates at fairwork.gov.au.',
       size=8, fg='7F0000', bg='FCE4D6', halign='left', wrap=True, bdr=False)
    ws.freeze_panes = 'A5'


def _casual_detail(ws, results, week_str):
    ws.sheet_view.showGridLines = False
    cas_results = [r for r in results if r['emp_type'] == 'Casual']

    ws.merge_cells('A1:Y1')
    ws.row_dimensions[1].height = 28
    sc(ws['A1'], f'CASUAL EMPLOYEE PAYROLL DETAIL  —  Week Ending {week_str}',
       bold=True, size=12, fg=HDR_FG, bg=HDR_BG, halign='center', valign='center')

    col_defs = [
        ('A', 22, 'Employee'), ('B', 7, 'Sect.'), ('C', 8, 'WD Hrs'), ('D', 10, 'WD Rate'),
        ('E', 11, 'WD Pay'), ('F', 8, 'Sat Hrs'), ('G', 10, 'Sat Rate'), ('H', 11, 'Sat Pay'),
        ('I', 8, 'Sun Hrs\n(≤2h)'), ('J', 10, 'Sun Rate'), ('K', 11, 'Sun Pay'),
        ('L', 9, 'Sun OT\nHrs'), ('M', 11, 'Sun OT\nPay'), ('N', 8, 'PH Hrs'),
        ('O', 10, 'PH Rate'), ('P', 11, 'PH Pay'), ('Q', 9, 'Dy OT Hrs'), ('R', 11, 'Dy OT Pay'),
        ('S', 9, 'Wk OT Hrs'), ('T', 11, 'Wk OT Pay'), ('U', 11, 'Laundry'),
        ('V', 8, 'LN Hrs'), ('W', 11, 'LN Pay'), ('X', 9, 'Total Hrs'), ('Y', 12, 'TOTAL PAY'),
    ]
    for col, width, _ in col_defs:
        ws.column_dimensions[col].width = width

    row = 2
    ws.row_dimensions[row].height = 24
    for col, _, hdr in col_defs:
        sc(ws[f'{col}{row}'], hdr, bold=True, size=9, fg=HDR_FG, bg='375623', halign='center', wrap=True)

    c_total_pay = c_total_hrs = 0.0
    for r_data in cas_results:
        row += 1
        ws.row_dimensions[row].height = 17
        hrs, pay, rates = r_data['hrs'], r_data['pay'], r_data['rates']
        dy_ot_hrs = hrs['daily_ot1'] + hrs['daily_ot2']
        dy_ot_pay = pay['daily_ot1'] + pay['daily_ot2']
        wk_ot_hrs = hrs['weekly_ot1'] + hrs['weekly_ot2']
        wk_ot_pay = pay['wk_ot1'] + pay['wk_ot2']
        row_vals = [
            (r_data['name'], None), (r_data['section'], None),
            (hrs['wd'], HRS_FMT), (rates['weekday'], CUR_FMT), (pay['wd'], CUR_FMT),
            (hrs['sat'], HRS_FMT), (rates['saturday'], CUR_FMT), (pay['sat'], CUR_FMT),
            (hrs['sun'], HRS_FMT), (rates['sunday'], CUR_FMT), (pay['sun'], CUR_FMT),
            (hrs['sun_ot'], HRS_FMT), (pay['sun_ot'], CUR_FMT),
            (hrs['ph'], HRS_FMT), (rates['ph_worked'], CUR_FMT), (pay['ph'], CUR_FMT),
            (dy_ot_hrs, HRS_FMT), (dy_ot_pay, CUR_FMT),
            (wk_ot_hrs, HRS_FMT), (wk_ot_pay, CUR_FMT),
            (pay['laundry'], CUR_FMT),
            (hrs['late_night'], HRS_FMT), (pay['late_night'], CUR_FMT),
            (hrs['total'], HRS_FMT), (pay['total'], CUR_FMT),
        ]
        bg_map = {2: CAS_BG, 3: CAS_BG, 4: CAS_BG, 5: SAT_BG, 6: SAT_BG, 7: SAT_BG,
                  8: SUN_BG, 9: SUN_BG, 10: SUN_BG, 11: OT_BG, 12: OT_BG,
                  13: PH_BG, 14: PH_BG, 15: PH_BG, 16: OT_BG, 17: OT_BG, 18: OT_BG, 19: OT_BG,
                  20: CAS_BG, 21: OT_BG, 22: OT_BG, 23: GREY_BG}
        cols_y = [c for c, _, _ in col_defs]
        for i, (val, fmt) in enumerate(row_vals):
            sc(ws[f'{cols_y[i]}{row}'], val, size=9, bg=bg_map.get(i, None), fmt=fmt,
               halign='center' if i > 1 else 'left', bold=(i == len(row_vals) - 1),
               fg=BLACK_FG if i > 0 else BLUE_FG)
        c_total_pay += pay['total']
        c_total_hrs += hrs['total']

    row += 1
    ws.row_dimensions[row].height = 18
    for col, _, _ in col_defs:
        c = ws[f'{col}{row}']
        if col == 'A':
            sc(c, 'TOTAL', bold=True, size=9, bg=TOTAL_BG)
        elif col == 'X':
            sc(c, c_total_hrs, bold=True, size=9, bg=TOTAL_BG, fmt=HRS_FMT, halign='center')
        elif col == 'Y':
            sc(c, c_total_pay, bold=True, size=9, bg=TOTAL_BG, fmt=CUR_FMT, halign='center')
        else:
            sc(c, '', bg=TOTAL_BG)
    ws.freeze_panes = 'A3'


def _labour_costs(ws, results, week_str):
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 28
    sc(ws['A1'], f'LABOUR COST REPORT  —  Week Ending {week_str}',
       bold=True, size=12, fg=HDR_FG, bg=HDR_BG, halign='center', valign='center')

    for col, w in [('A', 22), ('B', 14), ('C', 14), ('D', 14), ('E', 14), ('F', 14), ('G', 22)]:
        ws.column_dimensions[col].width = w

    row = 3
    ws.row_dimensions[row].height = 20
    for col, hdr in [('A', 'Section'), ('B', 'FT/PT Hours'), ('C', 'FT/PT Cost'),
                     ('D', 'Casual Hours'), ('E', 'Casual Cost'),
                     ('F', 'Total Hours'), ('G', 'Total Labour Cost')]:
        sc(ws[f'{col}{row}'], hdr, bold=True, size=9, fg=HDR_FG, bg=SUB_BG, halign='center', wrap=True)

    sections = {}
    for r in results:
        sec = r.get('section', 'Unknown') or 'Unknown'
        sections.setdefault(sec, {'perm_hrs': 0, 'perm_cost': 0, 'cas_hrs': 0, 'cas_cost': 0})
        if r['emp_type'] == 'Casual':
            sections[sec]['cas_hrs'] += r['hrs']['total']
            sections[sec]['cas_cost'] += r['award_pay']
        else:
            sections[sec]['perm_hrs'] += r['hrs']['total']
            sections[sec]['perm_cost'] += r['gross']

    bg_map = {'FOH': FOH_BG, 'BOH': BOH_BG}
    total_perm_hrs = total_perm_cost = total_cas_hrs = total_cas_cost = 0.0
    for sec, data in sorted(sections.items()):
        row += 1
        ws.row_dimensions[row].height = 18
        bg = bg_map.get(sec, GREY_BG)
        total_h = data['perm_hrs'] + data['cas_hrs']
        total_c = data['perm_cost'] + data['cas_cost']
        for col, val, fmt in [('A', sec, None), ('B', data['perm_hrs'], HRS_FMT),
                              ('C', data['perm_cost'], CUR_FMT), ('D', data['cas_hrs'], HRS_FMT),
                              ('E', data['cas_cost'], CUR_FMT), ('F', total_h, HRS_FMT),
                              ('G', total_c, CUR_FMT)]:
            sc(ws[f'{col}{row}'], val, size=10, bold=(col == 'G'), bg=bg, fmt=fmt,
               halign='center' if col != 'A' else 'left')
        total_perm_hrs += data['perm_hrs']; total_perm_cost += data['perm_cost']
        total_cas_hrs += data['cas_hrs']; total_cas_cost += data['cas_cost']

    row += 1
    ws.row_dimensions[row].height = 20
    grand_hrs = total_perm_hrs + total_cas_hrs
    grand_cost = total_perm_cost + total_cas_cost
    for col, val, fmt in [('A', 'TOTAL LABOUR COST', None), ('B', total_perm_hrs, HRS_FMT),
                          ('C', total_perm_cost, CUR_FMT), ('D', total_cas_hrs, HRS_FMT),
                          ('E', total_cas_cost, CUR_FMT), ('F', grand_hrs, HRS_FMT),
                          ('G', grand_cost, CUR_FMT)]:
        sc(ws[f'{col}{row}'], val, bold=True, size=10, bg=TOTAL_BG, fmt=fmt,
           halign='center' if col != 'A' else 'left')

    row += 2
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f'A{row}:G{row}')
    sc(ws[f'A{row}'], 'TOP-UP DETAIL  (amounts added to flat-rate employees to meet award minimum)',
       bold=True, size=9, fg=HDR_FG, bg='C55A11', halign='left')

    row += 1
    ws.row_dimensions[row].height = 20
    for col, hdr in [('A', 'Employee'), ('B', 'Section'), ('C', 'Employment'),
                     ('D', 'Flat Pay'), ('E', 'Award Pay'), ('F', 'TOP UP'), ('G', 'Gross Pay')]:
        sc(ws[f'{col}{row}'], hdr, bold=True, size=9, fg=HDR_FG, bg='7F3F00', halign='center')

    topup_results = [r for r in results if r.get('topup', 0) > 0.01]
    total_topup = 0.0
    for r in topup_results:
        row += 1
        ws.row_dimensions[row].height = 16
        for col, val, fmt in [('A', r['name'], None), ('B', r['section'], None),
                              ('C', r['emp_type'], None), ('D', r['flat_pay'], CUR_FMT),
                              ('E', r['award_pay'], CUR_FMT), ('F', r['topup'], CUR_FMT),
                              ('G', r['gross'], CUR_FMT)]:
            sc(ws[f'{col}{row}'], val, size=9, bg=PH_BG, fmt=fmt,
               halign='center' if col not in ('A', 'B', 'C') else 'left',
               bold=(col == 'F'), fg=RED_FG if col == 'F' else BLACK_FG)
        total_topup += r['topup']

    if not topup_results:
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        sc(ws[f'A{row}'], 'No top-ups required this week — all flat rates meet or exceed award.',
           size=9, bg=ZERO_BG, fg='375623', bold=True)
    else:
        row += 1
        ws.row_dimensions[row].height = 18
        for col, val, fmt in [('A', 'TOTAL TOP-UP', ''), ('B', '', ''), ('C', '', ''),
                              ('D', '', ''), ('E', '', ''), ('F', total_topup, CUR_FMT), ('G', '', '')]:
            sc(ws[f'{col}{row}'], val, bold=True, size=10, bg=TOTAL_BG, fmt=fmt,
               halign='center' if col != 'A' else 'left', fg=RED_FG if col == 'F' else BLACK_FG)


def _daily_breakdown(ws, results, week_str):
    ws.sheet_view.showGridLines = False
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 28
    sc(ws['A1'], f'DAILY BREAKDOWN  —  Week Ending {week_str}',
       bold=True, size=12, fg=HDR_FG, bg=HDR_BG, halign='center', valign='center')

    for col, w in [('A', 22), ('B', 8), ('C', 8), ('D', 11), ('E', 9),
                   ('F', 9), ('G', 9), ('H', 9), ('I', 9), ('J', 12)]:
        ws.column_dimensions[col].width = w

    row = 2
    ws.row_dimensions[row].height = 20
    for col, hdr in [('A', 'Employee'), ('B', 'Date'), ('C', 'Day'), ('D', 'Type'),
                     ('E', 'Shifts'), ('F', 'Total Hrs'), ('G', 'Ordinary'),
                     ('H', 'Daily OT'), ('I', 'Late Night'), ('J', 'Section')]:
        sc(ws[f'{col}{row}'], hdr, bold=True, size=9, fg=HDR_FG, bg=SUB_BG, halign='center')

    type_bg = {'wd': None, 'sat': SAT_BG, 'sun': SUN_BG, 'ph': PH_BG}
    type_label = {'wd': 'Weekday', 'sat': 'Saturday', 'sun': 'Sunday', 'ph': 'Public Hol.'}
    for r_data in results:
        for day in r_data.get('day_rows', []):
            row += 1
            ws.row_dimensions[row].height = 15
            dt = day['day_type']
            bg = type_bg.get(dt, None)
            dy_ot = day.get('daily_ot1', 0) + day.get('daily_ot2', 0)
            for col, val, fmt in [('A', r_data['name'], None),
                                  ('B', pd.Timestamp(day['date']).strftime('%d/%m'), None),
                                  ('C', day['day_name'], None), ('D', type_label.get(dt, dt), None),
                                  ('E', day['shifts'], '0'), ('F', day['total_hrs'], HRS_FMT),
                                  ('G', day['ordinary_hrs'], HRS_FMT), ('H', dy_ot, HRS_FMT),
                                  ('I', day['late_night'], HRS_FMT), ('J', r_data['section'], None)]:
                sc(ws[f'{col}{row}'], val, size=9, bg=bg, fmt=fmt,
                   halign='center' if col not in ('A', 'D', 'J') else 'left')
    ws.freeze_panes = 'A3'
