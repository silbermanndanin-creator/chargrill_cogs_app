"""Weekly Reconciliation tool — upload Tyro location reports, auto-fill the template.

Run:  .venv\\Scripts\\streamlit run reconcile_app.py   (or double-click "Reconcile Tool.bat")
"""
import io
import re
import datetime
from python_calamine import CalamineWorkbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment

# ----------------------------- styling -----------------------------
CUR = '"$"#,##0.00'
DATEFMT = 'dddd d mmm yyyy'
TEAL = '0F766E'; TEAL_D = '134E4A'; GOLD = 'FDE68A'; INPUT = 'FFF3CD'; WHITE = 'FFFFFF'

def fB(**k): return Font(name='Calibri', **k)
def fill(h): return PatternFill('solid', fgColor=h)
white_bold = fB(bold=True, color=WHITE)
blue = fB(color='0000FF'); blue_b = fB(bold=True, color='0000FF')
black = fB(color='000000'); green = fB(color='008000')
_thin = Side(style='thin', color='D0D0D0')
border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def put(ws, coord, val, font=None, numfmt=None, fillc=None, align=None, bd=False, comment=None):
    c = ws[coord]; c.value = val
    if font: c.font = font
    if numfmt: c.number_format = numfmt
    if fillc: c.fill = fill(fillc)
    if align: c.alignment = Alignment(horizontal=align, vertical='center')
    if bd: c.border = border
    if comment: c.comment = Comment(comment, "Reconciliation")
    return c

# ----------------------------- parsing -----------------------------
TERM_RE = re.compile(r'^\s*([123])\s*-\s*Terminal', re.I)

def parse_report(data_bytes):
    """Return [t1_net, t2_net, t3_net] from one Tyro location report (.xlsx bytes)."""
    wb = CalamineWorkbook.from_filelike(io.BytesIO(data_bytes))
    rows = wb.get_sheet_by_name(wb.sheet_names[0]).to_python()
    net_col = None
    for row in rows:
        for j, v in enumerate(row):
            if isinstance(v, str) and v.strip().lower() == 'net total':
                net_col = j
        if net_col is not None:
            break
    if net_col is None:
        net_col = -1  # fall back to last column
    nets = {}
    for row in rows:
        if not row:
            continue
        m = TERM_RE.match(str(row[0]))
        if m:
            try:
                nets[int(m.group(1))] = float(row[net_col])
            except (TypeError, ValueError):
                nets[int(m.group(1))] = None
    return [nets.get(1), nets.get(2), nets.get(3)]

def ts_key(name):
    """Sort key from the download timestamp in 'locationReport - 2026-05-25T121558.xlsx'."""
    m = re.search(r'T(\d{6})', name)
    return m.group(1) if m else name

# ----------------------------- workbook builder -----------------------------
DAY_ROWS = [7, 11, 15, 19, 23, 27, 31]

def build_workbook(week_start, tyro_days, deliv_days):
    """tyro_days/deliv_days: list of 7 dicts. Missing keys -> blank input cell.
    tyro keys: t1,t2,t3,pos,act,adj,turnover.  deliv keys: uber,doordash,bite."""
    wb = Workbook()
    ws = wb.active; ws.title = 'Tyro Net Totals'; ws.sheet_view.showGridLines = False
    for col, w in {'A':26,'B':14,'C':14,'D':14,'E':14,'F':7,'G':13,'H':13,'I':2,'J':16}.items():
        ws.column_dimensions[col].width = w

    put(ws, 'A1', 'Week commencing (Mon):', fB(bold=True), align='right')
    put(ws, 'B1', week_start, blue_b, DATEFMT, INPUT, 'left',
        comment='Monday of the week. All day labels update from this cell.')
    ws.merge_cells('A3:E4')
    t = ws['A3']; t.value = 'Terminal Net Totals Summary'
    t.font = fB(bold=True, color=WHITE, size=16); t.fill = fill(TEAL)
    t.alignment = Alignment(horizontal='center', vertical='center')
    for coord, txt in [('A6','Day'),('B6','Terminal 1 Net'),('C6','Terminal 2 Net'),
                       ('D6','Terminal 3 Net'),('E6','Daily Total')]:
        put(ws, coord, txt, white_bold, fillc=TEAL_D, align='center')

    for i, r in enumerate(DAY_ROWS):
        td = tyro_days[i] if i < len(tyro_days) else {}
        put(ws, f'A{r}', f'=$B$1+{i}', black, DATEFMT, bd=True)
        put(ws, f'B{r}', td.get('t1'), blue, CUR, INPUT, bd=True)
        put(ws, f'C{r}', td.get('t2'), blue, CUR, INPUT, bd=True)
        put(ws, f'D{r}', td.get('t3'), blue, CUR, INPUT, bd=True)
        put(ws, f'E{r}', f'=SUM(B{r}:D{r})', fB(bold=True), CUR, bd=True)
        if r == 7:
            put(ws, 'H7', 'Adjustments', fB(bold=True), align='center')
            put(ws, 'J7', 'Turnover (POS slip)', fB(bold=True), align='center')
        put(ws, f'F{r+1}', 'POS', fB(bold=True))
        put(ws, f'G{r+1}', td.get('pos'), blue, CUR, INPUT)
        put(ws, f'F{r+2}', 'ACT', fB(bold=True))
        put(ws, f'G{r+2}', td.get('act'), blue, CUR, INPUT)
        put(ws, f'H{r+2}', td.get('adj'), blue, CUR, INPUT)
        put(ws, f'J{r+2}', td.get('turnover'), blue, CUR, INPUT)

    put(ws, 'A35', 'Grand Total', fB(bold=True), fillc=GOLD)
    for col in ('B','C','D','E'):
        put(ws, f'{col}35', f'=SUM({col}7:{col}33)', fB(bold=True), CUR, GOLD)
    put(ws, 'F35', 'POS', fB(bold=True), fillc=GOLD)
    put(ws, 'G35', '=G8+G12+G16+G20+G24+G28+G32', fB(bold=True), CUR, GOLD)
    put(ws, 'F36', 'ACT', fB(bold=True))
    put(ws, 'G36', '=G9+G13+G17+G21+G25+G29+G33', fB(bold=True), CUR)
    put(ws, 'I35', 'Turnover', fB(bold=True), fillc=GOLD, align='right')
    put(ws, 'J35', '=SUM(J9:J33)', fB(bold=True), CUR, GOLD)
    put(ws, 'A37', 'Week cash check (ACT - POS + Adj):', fB(bold=True), align='right')
    put(ws, 'B37', '=G36-G35+SUM(H9,H13,H17,H21,H25,H29,H33)', fB(bold=True), CUR, align='center', bd=True,
        comment='Whole-week cash residual. Near 0 = good; red if off by more than $10.')
    ws.conditional_formatting.add('B37', FormulaRule(formula=['ABS(B37)>10'], fill=fill('F8D7DA'), font=Font(color='B02A37', bold=True)))
    ws.conditional_formatting.add('B37', FormulaRule(formula=['ABS(B37)<=10'], fill=fill('D1E7DD'), font=Font(color='0F5132', bold=True)))

    # ---- Deliveries & Bite ----
    d = wb.create_sheet('Deliveries & Bite'); d.sheet_view.showGridLines = False
    for col, w in {'A':26,'B':16,'C':16,'D':16,'E':2,'F':56}.items():
        d.column_dimensions[col].width = w
    d.merge_cells('A1:D2')
    h = d['A1']; h.value = 'Delivery Gross & Bite (App Payments)'
    h.font = fB(bold=True, color=WHITE, size=16); h.fill = fill(TEAL)
    h.alignment = Alignment(horizontal='center', vertical='center')
    put(d, 'A3', 'Week commencing (Mon):', fB(bold=True), align='right')
    b3 = d['B3']; b3.value = "='Tyro Net Totals'!$B$1"; b3.font = green; b3.number_format = DATEFMT
    for coord, txt in [('A5','Day'),('B5','Uber Eats gross'),('C5','DoorDash gross'),('D5','Bite (App pymt)')]:
        put(d, coord, txt, white_bold, fillc=TEAL_D, align='center')
    for i in range(7):
        r = 6 + i
        dd = deliv_days[i] if i < len(deliv_days) else {}
        put(d, f'A{r}', f'=$B$3+{i}', black, DATEFMT, bd=True)
        put(d, f'B{r}', dd.get('uber'), blue, CUR, INPUT, bd=True)
        put(d, f'C{r}', dd.get('doordash'), blue, CUR, INPUT, bd=True)
        put(d, f'D{r}', dd.get('bite'), blue, CUR, INPUT, bd=True)
    put(d, 'A13', 'Week Total', fB(bold=True), fillc=GOLD)
    for col in ('B','C','D'):
        put(d, f'{col}13', f'=SUM({col}6:{col}12)', fB(bold=True), CUR, GOLD)
    put(d, 'F5', 'Where each column goes in your existing files', fB(bold=True, size=12))
    for j, n in enumerate([
        'Uber Eats gross  ->  Uber.xlsx  >  UBER sheet  >  column C',
        'DoorDash gross   ->  Uber.xlsx  >  DOORDASH sheet  >  column C',
        'Bite (App pymt)  ->  App payments.xlsx  >  Sheet1  >  column C',
        '', 'Column A is the date in each file. Match the dates,',
        'then paste this column of 7 values down column C.']):
        if n: put(d, f'F{6+j}', n, fB(size=11))

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ----------------------------- Streamlit UI -----------------------------
def _run_ui():
    import streamlit as st
    import pandas as pd

    st.set_page_config(page_title="Weekly Reconciliation", page_icon="🧾", layout="wide")
    st.title("🧾 Weekly Reconciliation")
    st.caption("Upload the 7 Tyro location reports → auto-fill the terminal nets → add the rest → download the filled template.")

    # default to the most recent past Monday
    today = datetime.date.today()
    default_mon = today - datetime.timedelta(days=today.weekday() + 7)
    week_start = st.date_input("Week commencing (Monday)", value=default_mon)
    days = [week_start + datetime.timedelta(days=i) for i in range(7)]
    day_labels = [d.strftime("%a %d %b") for d in days]

    st.subheader("1 · Upload Tyro location reports")
    st.write("Download them **Monday → Sunday in order**; the tool sorts by download time. You can fix the order below.")
    files = st.file_uploader("Location report .xlsx files", type=["xlsx"], accept_multiple_files=True)

    tyro_nets = [[None, None, None] for _ in range(7)]
    if files:
        ordered = sorted(files, key=lambda f: ts_key(f.name))
        parsed = []
        for f in ordered:
            try:
                parsed.append((f.name, parse_report(f.getvalue())))
            except Exception as e:
                st.error(f"Could not read {f.name}: {e}")
        if len(parsed) != 7:
            st.warning(f"Got {len(parsed)} report(s); expected 7 (one per day). Fill any gaps manually below.")
        for i, (_, nets) in enumerate(parsed[:7]):
            tyro_nets[i] = nets

    st.subheader("2 · Terminal nets (auto-filled — edit if needed)")
    tyro_df = pd.DataFrame({
        "Day": day_labels,
        "Terminal 1 Net": [tyro_nets[i][0] for i in range(7)],
        "Terminal 2 Net": [tyro_nets[i][1] for i in range(7)],
        "Terminal 3 Net": [tyro_nets[i][2] for i in range(7)],
    })
    tyro_df = st.data_editor(tyro_df, hide_index=True, disabled=["Day"], key="tyro", width='stretch')
    tyro_df["Daily Total"] = tyro_df[["Terminal 1 Net", "Terminal 2 Net", "Terminal 3 Net"]].sum(axis=1, numeric_only=True)
    st.dataframe(tyro_df[["Day", "Daily Total"]], hide_index=True, width='stretch')

    st.subheader("3 · Cash + POS slip (type these in)")
    cash_df = st.data_editor(pd.DataFrame({
        "Day": day_labels, "POS": [None]*7, "ACT": [None]*7,
        "Adjustment": [None]*7, "Turnover (POS slip)": [None]*7,
    }), hide_index=True, disabled=["Day"], key="cash", width='stretch')

    st.subheader("4 · Deliveries + Bite (type these in)")
    deliv_df = st.data_editor(pd.DataFrame({
        "Day": day_labels, "Uber Eats gross": [None]*7,
        "DoorDash gross": [None]*7, "Bite (App pymt)": [None]*7,
    }), hide_index=True, disabled=["Day"], key="deliv", width='stretch')

    st.subheader("5 · Download")
    def num(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None
    tyro_days = [{
        "t1": num(tyro_df.iloc[i]["Terminal 1 Net"]), "t2": num(tyro_df.iloc[i]["Terminal 2 Net"]),
        "t3": num(tyro_df.iloc[i]["Terminal 3 Net"]), "pos": num(cash_df.iloc[i]["POS"]),
        "act": num(cash_df.iloc[i]["ACT"]), "adj": num(cash_df.iloc[i]["Adjustment"]),
        "turnover": num(cash_df.iloc[i]["Turnover (POS slip)"]),
    } for i in range(7)]
    deliv_days = [{
        "uber": num(deliv_df.iloc[i]["Uber Eats gross"]), "doordash": num(deliv_df.iloc[i]["DoorDash gross"]),
        "bite": num(deliv_df.iloc[i]["Bite (App pymt)"]),
    } for i in range(7)]
    buf = build_workbook(datetime.datetime.combine(week_start, datetime.time()), tyro_days, deliv_days)
    fname = f"{days[0].strftime('%d %b')} - {days[6].strftime('%d %b %Y')} Reconciliation.xlsx"
    st.download_button("⬇️ Download filled template", buf, file_name=fname,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    _run_ui()
