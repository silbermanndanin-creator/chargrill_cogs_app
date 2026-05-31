"""Food Safety Daily Temperature Records generator.

Builds a filled daily temperature record (one sheet per day) with randomised-but-
realistic values, following the venue's rules. Values are seeded by the date, so
regenerating a given day always produces the same sheet.
"""
import io
import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- styling ----
BLACK = '000000'; GREY = 'D9D9D9'; WHITE = 'FFFFFF'

def _f(**k): return Font(name='Calibri', **k)
bar_font = _f(bold=True, color=WHITE, size=11)
sub_font = _f(bold=True, size=9)
lbl_font = _f(size=10)
rng_font = _f(size=8, color='808080')
val_font = _f(size=11, color='1F3FBF')      # "entered" values, blue
title_font = _f(bold=True, size=14)
_thin = Side(style='thin', color='B0B0B0')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

MANAGERS = ["Danin", "Prakash", "Ashmita", "Mark", "Steve", "Jaskirat"]

# Which weekdays each delivery turns up (Mon=0 .. Sun=6). Rows on non-delivery
# days are left blank (no supplier, no temps).
DELIVERY_DAYS = {
    "Chilled Chicken": {0, 2, 4, 5},        # BAIDA: Mon, Wed, Fri, Sat
    "Chilled Salad": {0, 1, 2, 3, 4, 5},    # St George: Mon-Sat
    "Chilled Product": {0, 1, 2, 4, 5},     # BLUESEAS: Mon, Tue, Wed, Fri, Sat
}


def _fill(h): return PatternFill('solid', fgColor=h)

def _set(ws, coord, val, font=lbl_font, fill=None, align='left', border=True, wrap=False):
    c = ws[coord]
    c.value = val
    c.font = font
    if fill:
        c.fill = _fill(fill)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if border:
        c.border = BORDER
    return c

def _bar(ws, row, text, c0='A', c1='F'):
    ws.merge_cells(f'{c0}{row}:{c1}{row}')
    _set(ws, f'{c0}{row}', text, bar_font, BLACK, 'center')
    for col in _cols(c0, c1):
        ws[f'{col}{row}'].fill = _fill(BLACK)
        ws[f'{col}{row}'].border = BORDER

def _cols(c0, c1):
    return [chr(x) for x in range(ord(c0), ord(c1) + 1)]


def random_day(d):
    """Return a dict of all randomised values for date d (seeded by the date)."""
    rng = random.Random(d.toordinal())
    chilled = lambda: round(rng.uniform(2.3, 5.0), 1)          # mainly 2.3-5
    salad = lambda: round(rng.uniform(3.0, 5.0), 1)            # mainly 3-5
    hot = lambda: round(rng.uniform(75.0, 80.0), 1)           # hot bar 75-80
    cooked = lambda: round(rng.uniform(75.0, 83.0), 1)        # cooked 75-83
    mo, mc = rng.sample(MANAGERS, 2)

    # chicken cook records: 6 cooks, 8:30am then every 2h, +1h07 cook time
    def fmt(t):
        h, m = divmod(t, 60)
        ap = 'am' if h < 12 else 'pm'
        hh = h if 1 <= h <= 12 else (h - 12 if h > 12 else 12)
        return f"{hh}:{m:02d}{ap}"
    cooks = []
    for i in range(6):
        tin = 8 * 60 + 30 + i * 120
        cooks.append({"size": 6, "in": fmt(tin), "out": fmt(tin + 67),
                      "temp": round(rng.uniform(90.4, 94.8), 1)})

    return {
        "managers": (mo, mc),
        "deliveries": {  # only suppliers that deliver this weekday → (before 11am, before 5pm)
            name: (round(rng.uniform(2.5, 4.5), 1), round(rng.uniform(3.2, 5.0), 1))
            for name in ("Chilled Chicken", "Chilled Salad", "Chilled Product")
            if d.weekday() in DELIVERY_DAYS[name]},
        "products": {k: (chilled(), chilled()) for k in
                     ["Stuffed & Salted Ckn", "Peeled Chicken", "Short Cut Bacon", "Beef Patties"]},
        "hotbar": {  # devil wings & schnitzel: no 2hr-after-open value
            "Devil Wings": (None, hot()), "Schnitzel": (None, hot()),
            "Rice": (hot(), hot()), "Lamb": (hot(), hot())},
        "burger_chilled": {k: (chilled(), chilled()) for k in
                           ["Lettuce", "Tomato", "Cheese", "Chicken (RAW)", "Bacon (RAW)", "Beef Pattie (RAW)"]},
        "burger_cooked": {k: (cooked(), cooked()) for k in
                          ["Chicken Fillet (Cooked)", "Bacon (Cooked)", "Beef Pattie (Cooked)"]},
        "salad": {k: (salad(), salad()) for k in
                  ["Salad 1 (Right End)", "Salad 2 (Left End)", "Salad 3 (Left Middle)", "Salad 4 (Right Middle)"]},
        "desserts": {k: (salad(), salad()) for k in
                     ["Vanilla Cheesecake/Chocolate Mousse", "Pre Made Sauces", "Pre Made Soups"]},
        "equipment": {
            "Coolroom": (round(rng.uniform(1.8, 4.5), 1), round(rng.uniform(1.8, 4.5), 1)),
            "Salad Bar Fridge": (round(rng.uniform(1.8, 2.3), 1), round(rng.uniform(1.8, 2.3), 1)),
            "Display Fridge": (round(rng.uniform(2.4, 4.7), 1), round(rng.uniform(2.4, 4.7), 1)),
            "B/Station Fridge": (round(rng.uniform(1.3, 3.0), 1), round(rng.uniform(1.3, 3.0), 1)),
            "Hot Food Display": (rng.choice([79, 80]), rng.choice([79, 80])),
            "Clam Grill": (None, None),
            "Fryer 1": (186, 187), "Fryer 2": (180, 181), "Fryer 3": (164, 165)},
        "cooks": cooks,
    }


def build_day_sheet(wb, d, title=None):
    data = random_day(d)
    ws = wb.create_sheet(title or d.strftime("%Y-%m-%d"))
    ws.sheet_view.showGridLines = False
    for col, w in {'A': 24, 'B': 9, 'C': 11, 'D': 13, 'E': 10, 'F': 10,
                   'G': 2, 'H': 9, 'I': 11, 'J': 11, 'K': 12}.items():
        ws.column_dimensions[col].width = w

    def val(coord, v, align='center'):
        _set(ws, coord, v, val_font, align=align)

    # ---- title + header ----
    ws.merge_cells('A1:K1')
    _set(ws, 'A1', 'Food Safety Daily Temperature Records', title_font, align='center')
    _set(ws, 'A2', 'Day:', sub_font); val('B2', d.strftime('%A'), 'left')
    _set(ws, 'C2', 'Date:', sub_font); val('D2', d.strftime('%d/%m/%Y'), 'left')
    _set(ws, 'E2', 'Mgr Open:', sub_font); val('F2', data['managers'][0], 'left')
    _set(ws, 'H2', 'Mgr Close:', sub_font); ws.merge_cells('I2:J2'); val('I2', data['managers'][1], 'left')
    _set(ws, 'K2', '', border=True)

    # ---- Chicken Temp Records block (right side) ----
    _bar(ws, 4, 'Chicken Temp Records', 'H', 'K')
    for col, t in zip(['H', 'I', 'J', 'K'], ['Cook Size', 'Time In', 'Time Out', 'Temperature']):
        _set(ws, f'{col}5', t, sub_font, GREY, 'center')
    for i, ck in enumerate(data['cooks']):
        rr = 6 + i
        val(f'H{rr}', ck['size']); val(f'I{rr}', ck['in']); val(f'J{rr}', ck['out']); val(f'K{rr}', ck['temp'])

    r = 4
    # ---- Temperature of Deliveries ----
    _bar(ws, r, 'Temperature of Deliveries'); r += 1
    for col, t in zip('ABCDEF', ['Item', 'Temp', 'Supplier', 'Conditional Y/N', 'Before 11am', 'Before 5pm']):
        _set(ws, f'{col}{r}', t, sub_font, GREY, 'center')
    r += 1
    deliv = [('Chilled Chicken', '(1-5°C)', 'BAIDA'), ('Chilled Salad', '(1-5°C)', 'St George Food'),
             ('Chilled Product', '(1-5°C)', 'BLUESEAS'), ('Frozen', '(-15°C>)', '')]
    for name, rg, sup in deliv:
        _set(ws, f'A{r}', name); _set(ws, f'B{r}', rg, rng_font, align='center')
        _set(ws, f'D{r}', '')  # conditional column always blank
        if name in data['deliveries']:   # supplier delivered today
            _set(ws, f'C{r}', sup, val_font, align='center')
            b11, b5 = data['deliveries'][name]
            val(f'E{r}', b11); val(f'F{r}', b5)
        else:                            # no delivery this day → blank row
            _set(ws, f'C{r}', ''); _set(ws, f'E{r}', ''); _set(ws, f'F{r}', '')
        r += 1

    # ---- Shredded Chicken Process Time Log ----
    _bar(ws, r, 'Shredded Chicken Process Time Log'); r += 1
    for col, t in zip('ABCDEF', ['Time Completed', 'All bones removed', 'Team Member', 'Team Sign',
                                 'Shift Manager', 'Mgr Sign']):
        _set(ws, f'{col}{r}', t, sub_font, GREY, 'center')
    r += 1
    for col, v in zip('ABCDEF', ['3:00pm', 'Y', 'Jim', 'JN', 'Steve', 'SP']):
        val(f'{col}{r}', v)
    r += 1

    def section(title, rows, col_titles=('Temp', '2hrs after Open', '2hrs before Close', 'On Hand', 'Needed QTY')):
        nonlocal r
        _bar(ws, r, title); r += 1
        for col, t in zip('BCDEF', col_titles):
            _set(ws, f'{col}{r}', t, sub_font, GREY, 'center')
        _set(ws, f'A{r}', 'Product', sub_font, GREY, 'center')
        r += 1
        for name, rg, t1, t2 in rows:
            _set(ws, f'A{r}', name); _set(ws, f'B{r}', rg, rng_font, align='center')
            _set(ws, f'C{r}', '' if t1 is None else t1, val_font, align='center')
            _set(ws, f'D{r}', '' if t2 is None else t2, val_font, align='center')
            _set(ws, f'E{r}', ''); _set(ws, f'F{r}', '')
            r += 1

    section('Products and Holding Temperatures',
            [(k, '(1-5°C)', v[0], v[1]) for k, v in data['products'].items()])
    section('Hot Bar Holding Temperatures',
            [(k, '(60°C+)', ('n/a' if v[0] is None else v[0]), v[1]) for k, v in data['hotbar'].items()])
    section('Burger Station / Grilled Products Temperatures',
            [(k, '(1-5°C)', v[0], v[1]) for k, v in data['burger_chilled'].items()]
            + [(k, '(75°C+)', v[0], v[1]) for k, v in data['burger_cooked'].items()])
    section('Salad Bar Fridge Products Temperatures',
            [(k, '(1-5°C)', v[0], v[1]) for k, v in data['salad'].items()])
    section('Desserts / Sauces / Soups Temperature',
            [(k, '(1-5°C)', v[0], v[1]) for k, v in data['desserts'].items()])

    # ---- Equipment (two halves) ----
    _bar(ws, r, 'Equipment'); r += 1
    for col, t in zip('ABC', ['Unit/Item', '2hrs after Open', '2hrs before Close']):
        _set(ws, f'{col}{r}', t, sub_font, GREY, 'center')
    for col, t in zip('DEF', ['Unit/Item', '2hrs after Open', '2hrs before Close']):
        _set(ws, f'{col}{r}', t, sub_font, GREY, 'center')
    r += 1
    left = ["Coolroom", "Salad Bar Fridge", "Display Fridge", "B/Station Fridge"]
    right = ["Hot Food Display", "Clam Grill", "Fryer 1", "Fryer 2", "Fryer 3"]
    for i in range(max(len(left), len(right))):
        if i < len(left):
            k = left[i]; a, b = data['equipment'][k]
            _set(ws, f'A{r}', k); val(f'B{r}', a); val(f'C{r}', b)
        else:
            _set(ws, f'A{r}', ''); _set(ws, f'B{r}', ''); _set(ws, f'C{r}', '')
        if i < len(right):
            k = right[i]; a, b = data['equipment'][k]
            _set(ws, f'D{r}', k)
            _set(ws, f'E{r}', '' if a is None else a, val_font, align='center')
            _set(ws, f'F{r}', '' if b is None else b, val_font, align='center')
        else:
            _set(ws, f'D{r}', ''); _set(ws, f'E{r}', ''); _set(ws, f'F{r}', '')
        r += 1

    # ---- Daily Safety and Food Safety Checks ----
    _bar(ws, r, 'Daily Safety and Food Safety Checks'); r += 1
    _set(ws, f'A{r}', 'Check', sub_font, GREY, 'center')
    _set(ws, f'B{r}', 'Y/N', sub_font, GREY, 'center')
    ws.merge_cells(f'C{r}:F{r}'); _set(ws, f'C{r}', 'Notes', sub_font, GREY, 'center')
    r += 1
    checks = [
        ("Is the Thermometer being cleaned between raw and cooked?", "Y", ""),
        ("Are red cloths/chux being used in only raw areas?", "N", "Single use"),
        ("Is the Coolroom free of Mould build up? (Shelving, Fanguards)", "Y", ""),
        ("Is the floor clear, clean and dry, free from slip hazards?", "Y", ""),
        ("Are the Fire Exits free from obstructions?", "Y", ""),
        ("Are all team in full uniform, clean and tidy?", "Y", ""),
    ]
    for q, yn, note in checks:
        ws.merge_cells(f'A{r}:A{r}')
        _set(ws, f'A{r}', q, lbl_font, wrap=True)
        val(f'B{r}', yn)
        ws.merge_cells(f'C{r}:F{r}'); val(f'C{r}', note, 'left')
        r += 1

    ws.print_area = f'A1:K{r}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


def build_workbook(dates):
    wb = Workbook()
    wb.remove(wb.active)
    for d in dates:
        build_day_sheet(wb, d)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def build_single(d):
    return build_workbook([d])


def daterange(start, end):
    n = (end - start).days
    return [start + datetime.timedelta(days=i) for i in range(n + 1)]
